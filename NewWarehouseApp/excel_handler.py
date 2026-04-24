import os
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime

HEADER_FILL  = PatternFill("solid", fgColor="1976D2")
HEADER_FONT  = Font(bold=True, color="FFFFFF", size=11)
ALT_FILL     = PatternFill("solid", fgColor="F5F9FF")
THIN_BORDER  = Border(
    left=Side(style="thin", color="E0E0E0"),
    right=Side(style="thin", color="E0E0E0"),
    top=Side(style="thin", color="E0E0E0"),
    bottom=Side(style="thin", color="E0E0E0"),
)

# Maps every known header (Hebrew or English) in FinalOutput / pallet exports to a DB field
PALLET_HEADER_MAP = {
    'מק"ט':                                   "Pn",
    "חומר":                                   "Material",
    "יחידת מידה":                             "UnitOfMeasure",
    "סדרה":                                   "Batch",
    "WBS":                                    "WBS",
    "wbs":                                    "WBS",
    "אתר אחסון":                              "Storage",
    "סוג איחסון":                             "StorageType",
    "איתור איחסון":                           "Bin",
    "מלאי זמין":                              "Qty",
    "אסטרטגיה ייעודית":                       "DedicatedStrategy",
    "האם לפריט קיים יותר מאיתור יחיד":        "MultiLocation",
    "Pallet":                                 "PalletID",
    "Is_In_Stock":                            "IsInStock",
    "סביבתי / ממוזג":                         "Environment",
    "אזור WM":                                "WMArea",
    "איתור":                                  "TargetBin",
    # English export compat
    "PN":          "Pn",
    "Pn":          "Pn",
    "Batch":       "Batch",
    "Storage":     "Storage",
    "Area":        "StorageType",
    "Bin":         "Bin",
    "Qty":         "Qty",
    "IsInStock":   "IsInStock",
    "PalletID":    "PalletID",
    "Material":    "Material",
    "UnitOfMeasure": "UnitOfMeasure",
    "StorageType": "StorageType",
    "DedicatedStrategy": "DedicatedStrategy",
    "MultiLocation":     "MultiLocation",
    "Environment": "Environment",
    "WMArea":      "WMArea",
}


def _style_header(ws, headers):
    for col_idx, hdr in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=hdr)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = THIN_BORDER
    ws.row_dimensions[1].height = 28


def _auto_width(ws):
    for col in ws.columns:
        max_len = 0
        letter = col[0].column_letter
        for cell in col:
            try:
                max_len = max(max_len, len(str(cell.value or "")))
            except Exception:
                pass
        ws.column_dimensions[letter].width = min(max_len + 4, 40)


def load_pallet_xlsx(path: str, sheet: str = None) -> list[dict]:
    """Load FinalOutput.xlsx (Merge1 sheet) or any pallet export into row dicts."""
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb[sheet] if sheet and sheet in wb.sheetnames else wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        wb.close()
        return []

    raw_header = [str(h).strip() if h else "" for h in rows[0]]
    col_index: dict[str, int] = {}
    for i, h in enumerate(raw_header):
        db_field = PALLET_HEADER_MAP.get(h)
        if db_field and db_field not in col_index:
            col_index[db_field] = i

    result = []
    for row in rows[1:]:
        if all(v is None for v in row):
            continue
        d: dict = {}
        for field, idx in col_index.items():
            d[field] = str(row[idx]).strip() if (idx < len(row) and row[idx] is not None) else ""
        # Qty as float
        try:
            d["Qty"] = float(d.get("Qty", "") or 0)
        except ValueError:
            d["Qty"] = 0.0
        result.append(d)
    wb.close()
    return result


def load_locations_xlsx(path: str) -> list[dict]:
    """Load newBIns.xlsx (3-column: סביבתי / ממוזג, אזור WM, איתור)."""
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        wb.close()
        return []
    header = [str(h).strip() if h else "" for h in rows[0]]

    # Map Hebrew header → DB field
    LOC_MAP = {
        "סביבתי / ממוזג": "Environment",
        "אזור WM":         "DestArea",
        "איתור":           "Dest_BIN",
        # English fallback
        "DestArea":  "DestArea",
        "Dest_BIN":  "Dest_BIN",
        "Environment": "Environment",
    }

    col_index: dict[str, int] = {}
    for i, h in enumerate(header):
        db_field = LOC_MAP.get(h)
        if db_field and db_field not in col_index:
            col_index[db_field] = i

    result = []
    for row in rows[1:]:
        if all(v is None for v in row):
            continue
        d: dict = {
            "Environment": "",
            "DestArea":    "",
            "Dest_BIN":    "",
        }
        for field, idx in col_index.items():
            d[field] = str(row[idx]).strip() if (idx < len(row) and row[idx] is not None) else ""
        if d.get("DestArea") and d.get("Dest_BIN"):
            result.append(d)
    wb.close()
    return result


def export_pallets_xlsx(rows, path: str):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "PalletAssignments"
    headers = [
        "PalletID", "Status", "TargetBin", "AssignDate",
        'מק"ט', "חומר", "יחידת מידה", "סדרה", "WBS",
        "אתר אחסון", "סוג איחסון", "איתור איחסון", "מלאי זמין",
        "אסטרטגיה ייעודית", "האם > איתור אחד",
        "סביבתי / ממוזג", "אזור WM", "קיים פיזית",
    ]
    db_keys = [
        "PalletID", "Status", "TargetBin", "AssignDate",
        "Pn", "Material", "UnitOfMeasure", "Batch", "WBS",
        "Storage", "StorageType", "Bin", "Qty",
        "DedicatedStrategy", "MultiLocation",
        "Environment", "WMArea", "IsInStock",
    ]
    _style_header(ws, headers)
    for r_idx, row in enumerate(rows, 2):
        fill = ALT_FILL if r_idx % 2 == 0 else None
        for c_idx, key in enumerate(db_keys, 1):
            try:
                val = row[key]
            except (IndexError, KeyError):
                val = ""
            if key == "IsInStock":
                val = "כן" if val else "לא"
            cell = ws.cell(row=r_idx, column=c_idx, value=val)
            cell.border = THIN_BORDER
            if fill:
                cell.fill = fill
    _auto_width(ws)
    wb.save(path)


def load_pallet_codes_xlsx(path: str) -> list[str]:
    """Load a single-column Excel file (Pallets.xlsx) and return list of pallet ID strings."""
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    if not rows:
        return []
    codes = []
    for row in rows[1:]:  # skip header
        val = row[0] if row else None
        if val is not None:
            s = str(val).strip()
            if s:
                codes.append(s)
    return codes


def archive_path(archive_dir: str, prefix: str) -> str:
    os.makedirs(archive_dir, exist_ok=True)
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    return os.path.join(archive_dir, f"{prefix}_{ts}.xlsx")
