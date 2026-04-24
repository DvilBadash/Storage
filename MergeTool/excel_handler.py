import os
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime

HEADER_FILL = PatternFill("solid", fgColor="1976D2")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
ALT_FILL    = PatternFill("solid", fgColor="F5F9FF")
THIN_BORDER = Border(
    left=Side(style="thin", color="E0E0E0"), right=Side(style="thin", color="E0E0E0"),
    top=Side(style="thin", color="E0E0E0"),  bottom=Side(style="thin", color="E0E0E0"),
)

# Maps Hebrew STOCK.xlsx headers → StagingInventory DB fields
STOCK_HEADER_MAP = {
    'מק"ט':                                  "Pn",
    "חומר":                                  "Cat",
    "יחידת מידה":                            "UnitOfMeasure",
    "סדרה":                                  "Batch",
    "WBS":                                   "WBS",
    "wbs":                                   "WBS",
    "אתר אחסון":                             "Storage",
    "סוג איחסון":                            "Area",
    "איתור איחסון":                          "Bin",
    "מלאי זמין":                             "Qty",
    "אסטרטגיה ייעודית":                      "DestArea",
    "האם לפריט קיים יותר מאיתור יחיד":       "MultiLocation",
    # English fallbacks (from OldWarehouseApp export)
    "Cat":          "Cat",
    "Pn":           "Pn",
    "PN":           "Pn",
    "UnitOfMeasure": "UnitOfMeasure",
    "Batch":        "Batch",
    "WBS":          "WBS",
    "Storage":      "Storage",
    "Area":         "Area",
    "Bin":          "Bin",
    "DestArea":     "DestArea",
    "Qty":          "Qty",
    "MultiLocation": "MultiLocation",
    "PalletID":     "PalletID",
    "IsInStock":    "IsInStock",
    "AssignDate":   "AssignDate",
    "TargetBin":    "TargetBin",
}

STAGING_FIELDS = ["Cat", "Pn", "UnitOfMeasure", "Batch", "WBS", "Storage",
                  "Area", "Bin", "DestArea", "Qty", "MultiLocation",
                  "PalletID", "IsInStock", "AssignDate", "TargetBin"]


def _style_header(ws, headers):
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = THIN_BORDER
    ws.row_dimensions[1].height = 28


def _auto_width(ws):
    for col in ws.columns:
        max_len = max((len(str(cell.value or "")) for cell in col), default=8)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 42)


def load_old_warehouse_xlsx(path: str) -> list[dict]:
    """Load OldWarehouseApp InventoryExport OR STOCK.xlsx (Hebrew headers)."""
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        wb.close()
        return []

    raw_header = [str(h).strip() if h else "" for h in rows[0]]
    col_index: dict[str, int] = {}
    for i, h in enumerate(raw_header):
        db_field = STOCK_HEADER_MAP.get(h)
        if db_field and db_field not in col_index:
            col_index[db_field] = i

    result = []
    for row in rows[1:]:
        if all(v is None for v in row):
            continue
        d: dict = {f: "" for f in STAGING_FIELDS}
        for field, idx in col_index.items():
            if field in d:
                d[field] = str(row[idx]).strip() if (idx < len(row) and row[idx] is not None) else ""
        try:
            d["Qty"] = float(d["Qty"]) if d["Qty"] else 0.0
        except ValueError:
            d["Qty"] = 0.0
        try:
            d["PalletID"] = int(float(d["PalletID"])) if d["PalletID"] else None
        except ValueError:
            d["PalletID"] = None
        d["IsInStock"] = 1 if str(d.get("IsInStock", "")).strip() in ("1", "כן", "true", "True") else 0
        result.append(d)
    wb.close()
    return result


def load_new_warehouse_xlsx(path: str) -> list[dict]:
    """Load NewWarehouseApp PalletAssignments export file."""
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        wb.close()
        return []
    header = [str(h).strip() if h else "" for h in rows[0]]
    seen_pallets = set()
    result = []
    for row in rows[1:]:
        if all(v is None for v in row):
            continue
        d = {header[i]: (str(row[i]).strip() if row[i] is not None else "") for i in range(len(header))}
        pid_raw = d.get("PalletID", "")
        try:
            pid = int(float(pid_raw)) if pid_raw else None
        except ValueError:
            pid = None
        if pid and pid not in seen_pallets:
            seen_pallets.add(pid)
            result.append({
                "PalletID":   pid,
                "Status":     d.get("Status", ""),
                "TargetBin":  d.get("TargetBin", ""),
                "AssignDate": d.get("AssignDate", ""),
            })
    wb.close()
    return result


def export_unified_xlsx(rows, path: str):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "UnifiedInventory"
    headers = [
        "חומר", 'מק"ט', "יחידת מידה", "סדרה", "WBS",
        "אתר אחסון", "סוג איחסון", "איתור איחסון", "אסטרטגיה ייעודית",
        "מלאי זמין", "האם > איתור אחד",
        "PalletID", "קיים פיזית", "TargetBin", "סטטוס", "AssignDate", "MergeDate",
    ]
    db_keys = [
        "Cat", "Pn", "UnitOfMeasure", "Batch", "WBS",
        "Storage", "Area", "Bin", "DestArea",
        "Qty", "MultiLocation",
        "PalletID", "IsInStock", "TargetBin", "Status", "AssignDate", "MergeDate",
    ]
    _style_header(ws, headers)

    STATUS_COLORS = {"ממוקם": "C8E6C9", "יצא": "BBDEFB", "הוקם": "FFF9C4"}

    for r_idx, row in enumerate(rows, 2):
        status = row["Status"] or ""
        fill_color = STATUS_COLORS.get(status)
        fill = PatternFill("solid", fgColor=fill_color) if fill_color else (ALT_FILL if r_idx % 2 == 0 else None)
        for c_idx, key in enumerate(db_keys, 1):
            try:
                val = row[key]
            except (IndexError, KeyError):
                val = ""
            if key == "IsInStock":
                val = "כן" if val else "לא"
            cell = ws.cell(row=r_idx, column=c_idx, value=val)
            cell.border = THIN_BORDER
            if fill:
                cell.fill = fill
    _auto_width(ws)
    wb.save(path)


def archive_path(archive_dir: str, prefix: str) -> str:
    os.makedirs(archive_dir, exist_ok=True)
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    return os.path.join(archive_dir, f"{prefix}_{ts}.xlsx")
