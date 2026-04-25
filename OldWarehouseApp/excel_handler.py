import os
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime


# DB field names used throughout the app
INVENTORY_COLUMNS = ["Cat", "Pn", "UnitOfMeasure", "Batch", "WBS",
                     "Storage", "Area", "Bin", "DestArea", "Qty", "MultiLocation"]

# Maps every known Hebrew header from STOCK.xlsx to the DB field name
STOCK_HEADER_MAP = {
    'מק"ט':                                  "Pn",
    "חומר":                                  "Cat",
    "יחידת מידה":                            "UnitOfMeasure",
    "סדרה":                                  "Batch",
    "wbs":                                   "WBS",
    "WBS":                                   "WBS",
    "אתר אחסון":                             "Storage",
    "סוג איחסון":                            "Area",
    "איתור איחסון":                          "Bin",
    "מלאי זמין":                             "Qty",
    "אסטרטגיה ייעודית":                      "DestArea",
    "האם לפריט קיים יותר מאיתור יחיד":       "MultiLocation",
}

HEADER_FILL   = PatternFill("solid", fgColor="1976D2")
HEADER_FONT   = Font(bold=True, color="FFFFFF", size=11)
HEADER_ALIGN  = Alignment(horizontal="center", vertical="center")
ALT_FILL      = PatternFill("solid", fgColor="F5F9FF")
THIN_BORDER   = Border(
    left=Side(style="thin", color="E0E0E0"),
    right=Side(style="thin", color="E0E0E0"),
    top=Side(style="thin", color="E0E0E0"),
    bottom=Side(style="thin", color="E0E0E0"),
)


def _style_header(ws, headers: list):
    for col_idx, hdr in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=hdr)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGN
        cell.border = THIN_BORDER
    ws.row_dimensions[1].height = 28


def _auto_width(ws):
    for col in ws.columns:
        max_len = 0
        col_letter = col[0].column_letter
        for cell in col:
            try:
                max_len = max(max_len, len(str(cell.value or "")))
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = min(max_len + 4, 40)


def load_inventory_xlsx(path: str) -> list[dict]:
    """Load STOCK.xlsx (Hebrew headers) into InventoryOld row dicts."""
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        wb.close()
        return []

    raw_header = [str(h).strip() if h else "" for h in rows[0]]
    # Build index map: DB field → column index
    col_index: dict[str, int] = {}
    for i, h in enumerate(raw_header):
        db_field = STOCK_HEADER_MAP.get(h) or STOCK_HEADER_MAP.get(h.lower())
        if db_field and db_field not in col_index:
            col_index[db_field] = i
        # Also try direct English match (for backwards compat with old exports)
        if h in INVENTORY_COLUMNS and h not in col_index:
            col_index[h] = i

    result = []
    for row in rows[1:]:
        if all(v is None for v in row):
            continue
        d: dict = {}
        for field in INVENTORY_COLUMNS:
            idx = col_index.get(field)
            d[field] = str(row[idx]).strip() if (idx is not None and row[idx] is not None) else ""
        try:
            d["Qty"] = float(d["Qty"]) if d["Qty"] else 0.0
        except ValueError:
            d["Qty"] = 0.0
        result.append(d)
    wb.close()
    return result


def export_inventory_xlsx(rows, path: str):
    """Export InventoryOld with pallet assignments (English headers for inter-app use)."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "InventoryOld"
    all_cols = INVENTORY_COLUMNS + ["PalletID", "IsInStock", "AssignDate", "TargetBin"]
    _style_header(ws, all_cols)
    for r_idx, row in enumerate(rows, 2):
        fill = ALT_FILL if r_idx % 2 == 0 else None
        for c_idx, col in enumerate(all_cols, 1):
            val = row[col] if col in row.keys() else ""
            if col == "IsInStock":
                val = "כן" if val else "לא"
            cell = ws.cell(row=r_idx, column=c_idx, value=val)
            cell.border = THIN_BORDER
            if fill:
                cell.fill = fill
    _auto_width(ws)
    wb.save(path)


def export_pallet_xlsx(pallet_rows: list, path: str):
    """Export pallet list for NewWarehouseApp import."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Pallets"
    headers = ["PalletID", "CreateDate", "CreateUser", "Status",
               "PN", "Batch", "WBS", "Storage", "Area", "Bin",
               "UnitOfMeasure", "MultiLocation", "Qty", "IsInStock"]
    _style_header(ws, headers)
    for r_idx, row in enumerate(pallet_rows, 2):
        fill = ALT_FILL if r_idx % 2 == 0 else None
        vals = [row[h] if h in row.keys() else "" for h in headers]
        for c_idx, val in enumerate(vals, 1):
            cell = ws.cell(row=r_idx, column=c_idx, value=val)
            cell.border = THIN_BORDER
            if fill:
                cell.fill = fill
    _auto_width(ws)
    wb.save(path)


def load_pallet_import_xlsx(path: str) -> list[dict]:
    """Read a pallet import Excel.

    Accepts two formats:
    - Single-column with header "Pallet" or "PalletID" (simple pallet code list)
    - Multi-column export from OldWarehouseApp (PalletID + PN + Batch + ...)
    """
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        wb.close()
        return []
    # Normalise "Pallet" header → "PalletID" so both formats share the same key
    header = [
        "PalletID" if (str(h).strip() if h else "") in ("Pallet", "PalletID")
        else (str(h).strip() if h else "")
        for h in rows[0]
    ]
    result = []
    for row in rows[1:]:
        if all(v is None for v in row):
            continue
        d = {header[i]: (str(row[i]).strip() if row[i] is not None else "")
             for i in range(len(header)) if i < len(header)}
        if not d.get("PalletID", "").strip():
            continue
        result.append(d)
    wb.close()
    return result


def create_pallet_import_sample(path: str):
    """Create a sample pallet-import Excel showing the expected column layout."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Pallets"
    headers = ["PalletID", "CreateDate", "CreateUser", "Status",
               "PN", "Batch", "WBS", "Storage", "Area", "Bin",
               "UnitOfMeasure", "MultiLocation", "Qty", "IsInStock"]
    _style_header(ws, headers)
    sample_rows = [
        ["P-001", "", "", "הוקם", "MATERIAL-001", "BATCH-A", "WBS-001", "STORAGE-A", "AREA-1", "BIN-001", "יח", "לא", 10.0, "כן"],
        ["P-002", "", "", "הוקם", "MATERIAL-002", "BATCH-B", "WBS-002", "STORAGE-B", "AREA-2", "BIN-002", "יח", "לא",  5.0, "לא"],
    ]
    for r_idx, vals in enumerate(sample_rows, 2):
        for c_idx, val in enumerate(vals, 1):
            cell = ws.cell(row=r_idx, column=c_idx, value=val)
            cell.border = THIN_BORDER
    _auto_width(ws)
    wb.save(path)


def archive_path(archive_dir: str, prefix: str) -> str:
    os.makedirs(archive_dir, exist_ok=True)
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    return os.path.join(archive_dir, f"{prefix}_{ts}.xlsx")
