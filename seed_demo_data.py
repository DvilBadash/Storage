"""
Seed demo data into all three WMS databases and create Excel import files.
Run from the Storage root folder: python seed_demo_data.py
"""
import sys, os, sqlite3, random, string
from datetime import datetime, timedelta
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

random.seed(42)

BASE    = os.path.dirname(os.path.abspath(__file__))
OLD_DB  = os.path.join(BASE, "OldWarehouseApp",  "old_warehouse.db")
NEW_DB  = os.path.join(BASE, "NewWarehouseApp",  "new_warehouse.db")
MRG_DB  = os.path.join(BASE, "MergeTool",        "merge_tool.db")
OUT_DIR = os.path.join(BASE, "SampleData")
os.makedirs(OUT_DIR, exist_ok=True)

# ── Shared style helpers ──────────────────────────────────────────────────────
HDR_FILL   = PatternFill("solid", fgColor="1976D2")
HDR_FONT   = Font(bold=True, color="FFFFFF", size=11)
HDR_ALIGN  = Alignment(horizontal="center", vertical="center", wrap_text=True)
ALT_FILL   = PatternFill("solid", fgColor="F0F8FF")
THIN       = Border(
    left=Side(style="thin",color="CCCCCC"), right=Side(style="thin",color="CCCCCC"),
    top=Side(style="thin",color="CCCCCC"),  bottom=Side(style="thin",color="CCCCCC"),
)

def style_ws(ws, headers):
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.font, cell.fill, cell.alignment, cell.border = HDR_FONT, HDR_FILL, HDR_ALIGN, THIN
    ws.row_dimensions[1].height = 30
    ws.freeze_panes = "A2"

def auto_width(ws):
    for col in ws.columns:
        w = max((len(str(c.value or "")) for c in col), default=8)
        ws.column_dimensions[col[0].column_letter].width = min(w + 3, 38)

# ── Reference data ────────────────────────────────────────────────────────────
CATEGORIES = ["אלק", "מכנ", "חשמ", "פלס", "מתכ", "כימ", "גומ", "זכו", "עצ", "אחר"]

PART_PREFIXES = ["EL","ME","CH","PL","MT","RU","GL","WD","HY","PN",
                 "TR","CA","RE","DI","MO","SE","VA","PU","FI","CO"]

USERS_OLD = ["דוד כהן", "יוסי לוי", "רחל פרץ", "אבי גולן", "שרה ברק"]
USERS_NEW = ["מושה שפיר", "לימור אדר", "נתן כץ", "דינה בן-דוד", "אמיר רון"]
USERS_MRG = ["מנהל מערכת", "אנה פישר", "ירון עוז"]

# Old warehouse geography
OLD_STORAGES = {
    "מחסן-א": {
        "W41": [f"{r:02d}-{s:02d}-{b:02d}" for r in range(1,6) for s in range(1,8) for b in range(1,5)],
        "W42": [f"{r:02d}-{s:02d}-{b:02d}" for r in range(1,5) for s in range(1,6) for b in range(1,4)],
        "W43": [f"{r:02d}-{s:02d}-{b:02d}" for r in range(1,4) for s in range(1,5) for b in range(1,4)],
    },
    "מחסן-ב": {
        "AK10": [f"{r:02d}-{s:02d}-{b:02d}" for r in range(1,5) for s in range(1,7) for b in range(1,4)],
        "AK20": [f"{r:02d}-{s:02d}-{b:02d}" for r in range(1,4) for s in range(1,6) for b in range(1,5)],
    },
    "מחסן-ג": {
        "AL30": [f"{r:02d}-{s:02d}-{b:02d}" for r in range(1,5) for s in range(1,5) for b in range(1,4)],
        "AL31": [f"{r:02d}-{s:02d}-{b:02d}" for r in range(1,4) for s in range(1,4) for b in range(1,5)],
    },
    "מחסן-ד": {
        "AM30": [f"{r:02d}-{s:02d}-{b:02d}" for r in range(1,6) for s in range(1,5) for b in range(1,4)],
        "AN40": [f"{r:02d}-{s:02d}-{b:02d}" for r in range(1,5) for s in range(1,6) for b in range(1,4)],
    },
}

# New warehouse geography
DEST_AREAS = ["AJ10","AJ11","AJ12","AK20","AK21","AK22","AL30","AL31","AM40","AM41","AN50"]
NEW_BINS_PER_AREA = 30  # bins per dest area

DEST_BINS = {
    area: [f"B{i:02d}-{j:02d}" for i in range(1,7) for j in range(1,6)]
    for area in DEST_AREAS
}

WBS_CODES = [f"WBS-{x:03d}-{y:02d}" for x in range(1, 31) for y in range(1, 4)][:60]
BATCHES   = [f"B{random.randint(100000,999999)}" for _ in range(120)]

def rand_date(days_back=180):
    return (datetime.now() - timedelta(days=random.randint(0, days_back))).strftime("%Y-%m-%d %H:%M:%S")

# ── Generate 2500+ unique inventory rows ────────────────────────────────────
def generate_inventory_rows(count=2600):
    rows = []
    seen = set()
    # Build flat list of (storage, area, bin) combos
    location_pool = []
    for sto, areas in OLD_STORAGES.items():
        for area, bins in areas.items():
            for b in bins:
                location_pool.append((sto, area, b))
    random.shuffle(location_pool)

    pn_list = [f"{pfx}-{num:05d}" for pfx in PART_PREFIXES for num in range(1, 200)]
    random.shuffle(pn_list)

    dest_area_pool = DEST_AREAS + [None] * 5  # some items have no dest yet

    attempts = 0
    while len(rows) < count and attempts < count * 10:
        attempts += 1
        cat     = random.choice(CATEGORIES)
        pn      = random.choice(pn_list[:300])
        batch   = random.choice(BATCHES)
        wbs     = random.choice(WBS_CODES)
        loc     = random.choice(location_pool)
        storage, area, bin_ = loc
        key = (pn, batch, wbs, storage, area, bin_)
        if key in seen:
            continue
        seen.add(key)
        dest_area = random.choice(dest_area_pool)
        qty = round(random.uniform(1, 500), 2)
        rows.append({
            "Cat": cat, "Pn": pn, "Batch": batch, "WBS": wbs,
            "Storage": storage, "Area": area, "Bin": bin_,
            "DestArea": dest_area or "", "Qty": qty,
        })
    print(f"  Generated {len(rows)} unique inventory rows")
    return rows

# ── Database helpers ──────────────────────────────────────────────────────────
def conn(db_path):
    c = sqlite3.connect(db_path)
    c.row_factory = sqlite3.Row
    c.execute("PRAGMA foreign_keys = ON")
    return c

def seed_users(c, users, app_name):
    for name in users:
        c.execute("INSERT OR IGNORE INTO Users (FullName, IsActive) VALUES (?,1)", (name,))
    c.commit()
    print(f"  [{app_name}] {len(users)} users seeded")

def seed_log(c, users, app_name, actions):
    entries = []
    for _ in range(80):
        user   = random.choice(users)
        action = random.choice(actions)
        ts     = rand_date(60)
        detail = f"פעולת דמו: {action} ע\"י {user}"
        entries.append((ts, user, app_name, action, detail))
    c.executemany(
        "INSERT INTO LogEntries (Timestamp,UserName,ApplicationName,ActionType,Details) VALUES (?,?,?,?,?)",
        entries,
    )
    c.commit()
    print(f"  [{app_name}] {len(entries)} log entries seeded")

# ── Seed OldWarehouseApp ──────────────────────────────────────────────────────
def seed_old(inv_rows):
    print("\n[OldWarehouseApp] Seeding...")
    c = conn(OLD_DB)

    seed_users(c, USERS_OLD, "OldWarehouseApp")

    # Inventory
    c.executemany(
        "INSERT OR IGNORE INTO InventoryOld (Cat,Pn,Batch,WBS,Storage,Area,Bin,DestArea,Qty) "
        "VALUES (:Cat,:Pn,:Batch,:WBS,:Storage,:Area,:Bin,:DestArea,:Qty)",
        inv_rows,
    )
    c.commit()
    total_inv = c.execute("SELECT COUNT(*) FROM InventoryOld").fetchone()[0]
    print(f"  InventoryOld: {total_inv} rows")

    # Get inserted IDs
    all_ids = [r[0] for r in c.execute("SELECT InventoryID FROM InventoryOld ORDER BY InventoryID").fetchall()]
    random.shuffle(all_ids)
    ids_to_assign = all_ids[:1800]  # assign 1800 items to pallets

    # Create 60 pallets
    pallet_ids = list(range(1001, 1061))
    for pid in pallet_ids:
        user = random.choice(USERS_OLD)
        c.execute(
            "INSERT OR IGNORE INTO Pallets (PalletID, CreateDate, CreateUser, Status) VALUES (?,?,?,'הוקם')",
            (pid, rand_date(60), user),
        )
    c.commit()

    # Assign items to pallets
    chunk = len(ids_to_assign) // len(pallet_ids)
    assigned = 0
    for i, pid in enumerate(pallet_ids):
        chunk_ids = ids_to_assign[i * chunk : (i + 1) * chunk]
        for inv_id in chunk_ids:
            is_in = random.choices([0, 1], weights=[3, 7])[0]
            c.execute(
                "INSERT OR IGNORE INTO PALLET_Assignment (InventoryID,PalletID,IsInStock,AssignDate) "
                "VALUES (?,?,?,?)",
                (inv_id, pid, is_in, rand_date(30)),
            )
            assigned += 1
    c.commit()
    print(f"  Pallets: {len(pallet_ids)}  |  Assignments: {assigned}")

    # Update last_load_date
    c.execute("INSERT OR REPLACE INTO Settings VALUES ('last_load_date',?)",
              (datetime.now().strftime("%Y-%m-%d %H:%M:%S"),))
    c.commit()

    seed_log(c, USERS_OLD, "OldWarehouseApp",
             ["LOGIN","LOAD_INVENTORY_XLSX","ASSIGN_PALLET","CREATE_PALLET","UPDATE_ISINSTOCK","LOGOUT"])
    c.close()

# ── Seed NewWarehouseApp ──────────────────────────────────────────────────────
def seed_new():
    print("\n[NewWarehouseApp] Seeding...")
    c = conn(NEW_DB)

    seed_users(c, USERS_NEW, "NewWarehouseApp")

    # Locations
    loc_rows = []
    for area in DEST_AREAS:
        for b in DEST_BINS[area]:
            loc_rows.append((area, b))
    c.executemany("INSERT OR IGNORE INTO LocationNew (DestArea,Dest_BIN) VALUES (?,?)", loc_rows)
    c.commit()
    print(f"  LocationNew: {len(loc_rows)} rows")

    # Pallets (mirror of old warehouse)
    pallet_ids = list(range(1001, 1061))
    statuses   = ["הוקם"] * 15 + ["ממוקם"] * 35 + ["יצא"] * 10
    random.shuffle(statuses)

    for i, pid in enumerate(pallet_ids):
        status = statuses[i]
        area   = random.choice(DEST_AREAS)
        bin_   = random.choice(DEST_BINS[area])
        target = f"{area}/{bin_}" if status in ("ממוקם", "יצא") else None
        assign_date = rand_date(20) if target else None
        user   = random.choice(USERS_NEW)
        c.execute(
            "INSERT OR IGNORE INTO Pallets (PalletID,CreateDate,CreateUser,Status,TargetBin,AssignDate) "
            "VALUES (?,?,?,?,?,?)",
            (pid, rand_date(60), user, status, target, assign_date),
        )

    # PalletItems
    c_old   = conn(OLD_DB)
    pa_rows = c_old.execute(
        "SELECT pa.PalletID, i.Pn, i.Batch, i.WBS, i.Storage, i.Area, i.Bin, i.Qty, pa.IsInStock "
        "FROM PALLET_Assignment pa JOIN InventoryOld i ON pa.InventoryID=i.InventoryID"
    ).fetchall()
    c_old.close()

    item_data = [(r["PalletID"], r["Pn"], r["Batch"], r["WBS"],
                  r["Storage"], r["Area"], r["Bin"], r["Qty"], r["IsInStock"]) for r in pa_rows]
    c.executemany(
        "INSERT INTO PalletItems (PalletID,Pn,Batch,WBS,Storage,Area,Bin,Qty,IsInStock) "
        "VALUES (?,?,?,?,?,?,?,?,?)",
        item_data,
    )
    c.commit()
    print(f"  Pallets: {len(pallet_ids)}  |  PalletItems: {len(item_data)}")

    c.execute("INSERT OR REPLACE INTO Settings VALUES ('last_pallet_import',?)",
              (datetime.now().strftime("%Y-%m-%d %H:%M:%S"),))
    c.commit()

    seed_log(c, USERS_NEW, "NewWarehouseApp",
             ["LOGIN","IMPORT_PALLETS","LOAD_LOCATIONS","ASSIGN_PALLET_BIN",
              "UPDATE_PALLET_STATUS","EXPORT_ASSIGNMENTS","LOGOUT"])
    c.close()

# ── Seed MergeTool ────────────────────────────────────────────────────────────
def seed_merge(inv_rows):
    print("\n[MergeTool] Seeding...")
    c = conn(MRG_DB)

    seed_users(c, USERS_MRG, "MergeTool")

    # Populate StagingInventory from inv_rows + pallet assignments
    c_old  = conn(OLD_DB)
    pa_map = {}
    for row in c_old.execute(
        "SELECT pa.InventoryID, pa.PalletID, pa.IsInStock, pa.AssignDate, pa.TargetBin "
        "FROM PALLET_Assignment pa"
    ).fetchall():
        pa_map[row["InventoryID"]] = row

    staging = []
    for row in c_old.execute("SELECT * FROM InventoryOld").fetchall():
        pa = pa_map.get(row["InventoryID"])
        staging.append((
            row["Cat"], row["Pn"], row["Batch"], row["WBS"],
            row["Storage"], row["Area"], row["Bin"], row["DestArea"], row["Qty"],
            pa["PalletID"]   if pa else None,
            pa["IsInStock"]  if pa else 0,
            pa["AssignDate"] if pa else None,
            pa["TargetBin"]  if pa else None,
        ))
    c_old.close()

    c.executemany(
        "INSERT INTO StagingInventory "
        "(Cat,Pn,Batch,WBS,Storage,Area,Bin,DestArea,Qty,PalletID,IsInStock,AssignDate,TargetBin) "
        "VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?)",
        staging,
    )

    # StagingNewWarehouse
    c_new = conn(NEW_DB)
    for p in c_new.execute("SELECT * FROM Pallets").fetchall():
        c.execute(
            "INSERT INTO StagingNewWarehouse (PalletID,Status,TargetBin,AssignDate) VALUES (?,?,?,?)",
            (p["PalletID"], p["Status"], p["TargetBin"], p["AssignDate"]),
        )
    c_new.close()
    c.commit()
    print(f"  StagingInventory: {len(staging)} rows loaded")

    # Run merge to populate UnifiedInventory
    from MergeTool.database import run_merge as mt_merge
    try:
        # Direct SQL merge since we can't import easily
        now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        new_ws = {}
        for row in c.execute("SELECT * FROM StagingNewWarehouse").fetchall():
            pid = row["PalletID"]
            if pid:
                new_ws[int(pid)] = {"Status": row["Status"] or "הוקם",
                                    "TargetBin": row["TargetBin"] or "",
                                    "AssignDate": row["AssignDate"] or ""}
        c.execute("DELETE FROM UnifiedInventory")
        for row in c.execute("SELECT * FROM StagingInventory").fetchall():
            pid = row["PalletID"]
            target_bin  = row["TargetBin"] or ""
            status      = "הוקם"
            assign_date = row["AssignDate"] or ""
            if pid:
                pid = int(pid)
                if pid in new_ws:
                    nw = new_ws[pid]
                    target_bin  = nw["TargetBin"] or target_bin
                    status      = nw["Status"]
                    assign_date = nw["AssignDate"] or assign_date
            c.execute(
                "INSERT INTO UnifiedInventory "
                "(Cat,Pn,Batch,WBS,Storage,Area,Bin,DestArea,Qty,"
                " PalletID,IsInStock,TargetBin,Status,AssignDate,MergeDate) "
                "VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)",
                (row["Cat"],row["Pn"],row["Batch"],row["WBS"],row["Storage"],
                 row["Area"],row["Bin"],row["DestArea"],row["Qty"],
                 pid if pid else None, row["IsInStock"],
                 target_bin, status, assign_date, now),
            )
        c.commit()
        cnt = c.execute("SELECT COUNT(*) FROM UnifiedInventory").fetchone()[0]
        print(f"  UnifiedInventory: {cnt} rows")
    except Exception as e:
        print(f"  Merge inline error: {e}")

    c.execute("INSERT OR REPLACE INTO Settings VALUES ('last_merge_date',?)",
              (datetime.now().strftime("%Y-%m-%d %H:%M:%S"),))
    c.commit()

    seed_log(c, USERS_MRG, "MergeTool",
             ["LOGIN","LOAD_OLD_FILE","LOAD_NEW_FILE","MERGE","EXPORT_UNIFIED","LOGOUT"])
    c.close()

# ── Excel export helpers ──────────────────────────────────────────────────────
def write_xl(path, sheet_name, headers, rows_data):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_name
    style_ws(ws, headers)
    for r_idx, row in enumerate(rows_data, 2):
        fill = ALT_FILL if r_idx % 2 == 0 else None
        for c_idx, val in enumerate(row, 1):
            cell = ws.cell(row=r_idx, column=c_idx, value=val)
            cell.border = THIN
            cell.alignment = Alignment(vertical="center")
            if fill:
                cell.fill = fill
        ws.row_dimensions[r_idx].height = 18
    auto_width(ws)
    wb.save(path)
    print(f"  Created: {os.path.basename(path)}  ({len(rows_data)} rows)")

# ── Create Excel import files ─────────────────────────────────────────────────
def create_excel_files(inv_rows):
    print("\n[Excel Files] Creating in SampleData/...")

    # 1. InventoryOld import file (2600 rows)
    inv_headers = ["Cat","Pn","Batch","WBS","Storage","Area","Bin","DestArea","Qty"]
    inv_data    = [(r["Cat"],r["Pn"],r["Batch"],r["WBS"],r["Storage"],
                    r["Area"],r["Bin"],r["DestArea"],r["Qty"]) for r in inv_rows]
    write_xl(os.path.join(OUT_DIR, "InventoryOld_Import.xlsx"),
             "InventoryOld", inv_headers, inv_data)

    # 2. LocationNew import file
    loc_headers = ["DestArea", "Dest_BIN"]
    loc_data    = [(area, b) for area in DEST_AREAS for b in DEST_BINS[area]]
    write_xl(os.path.join(OUT_DIR, "LocationNew_Import.xlsx"),
             "LocationNew", loc_headers, loc_data)

    # 3. PalletExport (from OldWarehouseApp → import into NewWarehouseApp)
    c_old = conn(OLD_DB)
    pal_rows = c_old.execute("""
        SELECT p.PalletID, p.CreateDate, p.CreateUser, p.Status,
               i.Pn AS PN, i.Batch, i.WBS, i.Storage, i.Area, i.Bin, i.Qty,
               pa.IsInStock
        FROM PALLET_Assignment pa
        JOIN InventoryOld i ON pa.InventoryID = i.InventoryID
        JOIN Pallets p       ON pa.PalletID   = p.PalletID
        ORDER BY p.PalletID, i.Pn
    """).fetchall()
    c_old.close()
    pal_headers = ["PalletID","CreateDate","CreateUser","Status",
                   "PN","Batch","WBS","Storage","Area","Bin","Qty","IsInStock"]
    pal_data = [(r["PalletID"],r["CreateDate"],r["CreateUser"],r["Status"],
                 r["PN"],r["Batch"],r["WBS"],r["Storage"],r["Area"],r["Bin"],
                 r["Qty"], 1 if r["IsInStock"] else 0) for r in pal_rows]
    write_xl(os.path.join(OUT_DIR, "PalletExport_OldWarehouse.xlsx"),
             "PalletExport", pal_headers, pal_data)

    # 4. NewWarehouse assignments export (for MergeTool)
    c_new = conn(NEW_DB)
    new_rows = c_new.execute("""
        SELECT p.PalletID, p.Status, p.TargetBin, p.AssignDate,
               pi.Pn, pi.Batch, pi.WBS, pi.Storage, pi.Area, pi.Bin, pi.Qty
        FROM Pallets p
        LEFT JOIN PalletItems pi ON p.PalletID = pi.PalletID
        ORDER BY p.PalletID
    """).fetchall()
    c_new.close()
    new_headers = ["PalletID","Status","TargetBin","AssignDate",
                   "PN","Batch","WBS","Storage","Area","Bin","Qty"]
    new_data = [(r["PalletID"],r["Status"],r["TargetBin"] or "",r["AssignDate"] or "",
                 r["Pn"],r["Batch"],r["WBS"],r["Storage"],r["Area"],r["Bin"],r["Qty"])
                for r in new_rows]
    write_xl(os.path.join(OUT_DIR, "PalletAssignments_NewWarehouse.xlsx"),
             "PalletAssignments", new_headers, new_data)

    # 5. UnifiedInventory (MergeTool output sample)
    c_mrg = conn(MRG_DB)
    uni_rows = c_mrg.execute("SELECT * FROM UnifiedInventory ORDER BY Pn, Batch").fetchall()
    c_mrg.close()
    uni_headers = ["Cat","PN","Batch","WBS","Storage","Area","Bin","DestArea","Qty",
                   "PalletID","IsInStock","TargetBin","Status","AssignDate","MergeDate"]
    uni_data = [(r["Cat"],r["Pn"],r["Batch"],r["WBS"],r["Storage"],r["Area"],r["Bin"],
                 r["DestArea"],r["Qty"],r["PalletID"],"כן" if r["IsInStock"] else "לא",
                 r["TargetBin"],r["Status"],r["AssignDate"],r["MergeDate"])
                for r in uni_rows]
    write_xl(os.path.join(OUT_DIR, "UnifiedInventory_Sample.xlsx"),
             "UnifiedInventory", uni_headers, uni_data)

# ── Main ──────────────────────────────────────────────────────────────────────
if __name__ == "__main__":
    print("=" * 60)
    print(" WMS Demo Data Seeder")
    print("=" * 60)

    # Init DBs first (runs schema creation)
    for db_path, app_dir in [(OLD_DB, "OldWarehouseApp"),
                              (NEW_DB, "NewWarehouseApp"),
                              (MRG_DB, "MergeTool")]:
        sys.path.insert(0, os.path.join(BASE, app_dir))
        import importlib
        db_mod = importlib.import_module("database")
        db_mod.init_db()
        sys.path.pop(0)
    print("\nAll databases initialized.")

    # Generate inventory
    print("\n[Inventory] Generating 2600+ unique rows...")
    inv_rows = generate_inventory_rows(2600)

    seed_old(inv_rows)
    seed_new()
    seed_merge(inv_rows)
    create_excel_files(inv_rows)

    # Summary
    print("\n" + "=" * 60)
    print(" Summary")
    print("=" * 60)
    for label, db_path, queries in [
        ("OldWarehouseApp", OLD_DB, [
            ("Users",           "SELECT COUNT(*) FROM Users"),
            ("InventoryOld",    "SELECT COUNT(*) FROM InventoryOld"),
            ("Pallets",         "SELECT COUNT(*) FROM Pallets"),
            ("PALLET_Assignment","SELECT COUNT(*) FROM PALLET_Assignment"),
            ("LogEntries",      "SELECT COUNT(*) FROM LogEntries"),
        ]),
        ("NewWarehouseApp", NEW_DB, [
            ("Users",       "SELECT COUNT(*) FROM Users"),
            ("LocationNew", "SELECT COUNT(*) FROM LocationNew"),
            ("Pallets",     "SELECT COUNT(*) FROM Pallets"),
            ("PalletItems", "SELECT COUNT(*) FROM PalletItems"),
            ("LogEntries",  "SELECT COUNT(*) FROM LogEntries"),
        ]),
        ("MergeTool", MRG_DB, [
            ("UnifiedInventory",    "SELECT COUNT(*) FROM UnifiedInventory"),
            ("StagingInventory",    "SELECT COUNT(*) FROM StagingInventory"),
            ("StagingNewWarehouse", "SELECT COUNT(*) FROM StagingNewWarehouse"),
            ("LogEntries",          "SELECT COUNT(*) FROM LogEntries"),
        ]),
    ]:
        print(f"\n  {label}:")
        c = conn(db_path)
        for name, sql in queries:
            cnt = c.execute(sql).fetchone()[0]
            print(f"    {name:<25} {cnt:>6} rows")
        c.close()

    print(f"\n  Excel files in: {OUT_DIR}")
    for f in sorted(os.listdir(OUT_DIR)):
        size = os.path.getsize(os.path.join(OUT_DIR, f)) // 1024
        print(f"    {f:<45} {size:>5} KB")

    print("\n Done! All demo data seeded successfully.")
